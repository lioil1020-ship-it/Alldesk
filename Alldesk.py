import tkinter as tk
import subprocess
import os
import stat
import sys
import tempfile
import time
import ctypes
import winreg
import shutil
import uuid

# third-party / local imports (optional)
try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

try:
    import win32com.client as win32com_client
except Exception:
    win32com_client = None

try:
    from pywinauto import Application as PywinautoApplication
except Exception:
    PywinautoApplication = None

from pathlib import Path
from tkinter import font as tkfont
from tkinter import ttk
from tkinter import messagebox
from urllib.parse import quote


def log_and_show(title: str, msg: str, level: str = 'warning'):
    """簡單的 log + 顯示 helper。level 可以是 'info'/'warning'/'error'。"""
    # console logging removed to avoid terminal output
    try:
        if level == 'error':
            messagebox.showerror(title, msg)
        elif level == 'info':
            messagebox.showinfo(title, msg)
        else:
            messagebox.showwarning(title, msg)
    except Exception:
        pass


# debug logging removed per user request: delete function and all call sites

 
# 輕量 DES 實作(支援單一 8-byte 區塊的 ECB 加密)
# 提供兼容介面:DES.new(key, DES.MODE_ECB).encrypt(data)
class _DES:
    """
    簡易 DES 實作(支援單一 8-byte 區塊的 ECB 加密).

    提供內部 API,模擬 Crypto 庫的行為,使呼叫端可以使用
    `DES.new(key, DES.MODE_ECB).encrypt(data)` 的介面.
    此實作僅用於相容性與小型工具,不建議用於生產環境.
    """

    def __init__(self, key: bytes):
        """初始化 DES 實例.

        參數:
        - key: 8 位元組的金鑰 (bytes)

        例外:
        - TypeError: key 非 bytes-like
        - ValueError: key 長度非 8
        """
        if not isinstance(key, (bytes, bytearray)):
            raise TypeError('key must be bytes')
        if len(key) != 8:
            raise ValueError('DES key must be 8 bytes')
        self.key = bytes(key)
        self.subkeys = self._generate_subkeys(self.key)

    @staticmethod
    def _bytes_to_bits(b: bytes):
        """將 bytes 轉為位元列表(MSB first).

        輸入例: b"\x01" -> [0,0,0,0,0,0,0,1]
        """
        bits = []
        for byte in b:
            for i in range(8)[::-1]:
                bits.append((byte >> i) & 1)
        return bits

    @staticmethod
    def _bits_to_bytes(bits):
        """將位元列表 (MSB first) 轉回 bytes.

        只支援位元數為 8 的整數倍.
        """
        out = bytearray()
        for i in range(0, len(bits), 8):
            byte = 0
            for bit in bits[i:i+8]:
                byte = (byte << 1) | bit
            out.append(byte)
        return bytes(out)

    def _permute(self, table, bits):
        """依照 permutation 表對位元列表重新排列並回傳新列表.

        table 為 1-based 的索引表.
        """
        return [bits[i-1] for i in table]

    def _left_rotate(self, lst, n):
        """將序列向左旋轉 n 位元.

        用於 DES subkey 的 C/D bits 旋轉.
        """
        return lst[n:]+lst[:n]

    def _generate_subkeys(self, key8: bytes):
        """從 8-byte 原始金鑰產生 16 個 48-bit 子金鑰.

        回傳值為 list[list[int]],每個子金鑰為位元 (0/1) 列表.
        """
        # PC-1
        pc1 = [57,49,41,33,25,17,9,
               1,58,50,42,34,26,18,
               10,2,59,51,43,35,27,
               19,11,3,60,52,44,36,
               63,55,47,39,31,23,15,
               7,62,54,46,38,30,22,
               14,6,61,53,45,37,29,
               21,13,5,28,20,12,4]

        # PC-2
        pc2 = [14,17,11,24,1,5,
               3,28,15,6,21,10,
               23,19,12,4,26,8,
               16,7,27,20,13,2,
               41,52,31,37,47,55,
               30,40,51,45,33,48,
               44,49,39,56,34,53,
               46,42,50,36,29,32]

        # rotations
        rotations = [1,1,2,2,2,2,2,2,1,2,2,2,2,2,2,1]

        key_bits = self._bytes_to_bits(key8)
        permuted = self._permute(pc1, key_bits)
        c = permuted[:28]
        d = permuted[28:]
        subkeys = []
        for r in rotations:
            c = self._left_rotate(c, r)
            d = self._left_rotate(d, r)
            cd = c + d
            sub = self._permute(pc2, cd)
            subkeys.append(sub)
        return subkeys

    def _feistel(self, r, subkey):
        """DES 的 Feistel 函數 (f 函數).

        參數:
        - r: 右半部位元列表 (32 bits)
        - subkey: 本輪的子金鑰 (48 bits)

        回傳 32-bit 的位元列表.
        """
        # Expansion table
        e_table = [32,1,2,3,4,5,4,5,6,7,8,9,8,9,10,11,12,13,12,13,14,15,16,17,16,17,18,19,20,21,20,21,22,23,24,25,24,25,26,27,28,29,28,29,30,31,32,1]

        # S-boxes
        s_boxes = [
            # S1
            [
                14,4,13,1,2,15,11,8,3,10,6,12,5,9,0,7,
                0,15,7,4,14,2,13,1,10,6,12,11,9,5,3,8,
                4,1,14,8,13,6,2,11,15,12,9,7,3,10,5,0,
                15,12,8,2,4,9,1,7,5,11,3,14,10,0,6,13
            ],
            # S2
            [
                15,1,8,14,6,11,3,4,9,7,2,13,12,0,5,10,
                3,13,4,7,15,2,8,14,12,0,1,10,6,9,11,5,
                0,14,7,11,10,4,13,1,5,8,12,6,9,3,2,15,
                13,8,10,1,3,15,4,2,11,6,7,12,0,5,14,9
            ],
            # S3
            [
                10,0,9,14,6,3,15,5,1,13,12,7,11,4,2,8,
                13,7,0,9,3,4,6,10,2,8,5,14,12,11,15,1,
                13,6,4,9,8,15,3,0,11,1,2,12,5,10,14,7,
                1,10,13,0,6,9,8,7,4,15,14,3,11,5,2,12
            ],
            # S4
            [
                7,13,14,3,0,6,9,10,1,2,8,5,11,12,4,15,
                13,8,11,5,6,15,0,3,4,7,2,12,1,10,14,9,
                10,6,9,0,12,11,7,13,15,1,3,14,5,2,8,4,
                3,15,0,6,10,1,13,8,9,4,5,11,12,7,2,14
            ],
            # S5
            [
                2,12,4,1,7,10,11,6,8,5,3,15,13,0,14,9,
                14,11,2,12,4,7,13,1,5,0,15,10,3,9,8,6,
                4,2,1,11,10,13,7,8,15,9,12,5,6,3,0,14,
                11,8,12,7,1,14,2,13,6,15,0,9,10,4,5,3
            ],
            # S6
            [
                12,1,10,15,9,2,6,8,0,13,3,4,14,7,5,11,
                10,15,4,2,7,12,9,5,6,1,13,14,0,11,3,8,
                9,14,15,5,2,8,12,3,7,0,4,10,1,13,11,6,
                4,3,2,12,9,5,15,10,11,14,1,7,6,0,8,13
            ],
            # S7
            [
                4,11,2,14,15,0,8,13,3,12,9,7,5,10,6,1,
                13,0,11,7,4,9,1,10,14,3,5,12,2,15,8,6,
                1,4,11,13,12,3,7,14,10,15,6,8,0,5,9,2,
                6,11,13,8,1,4,10,7,9,5,0,15,14,2,3,12
            ],
            # S8
            [
                13,2,8,4,6,15,11,1,10,9,3,14,5,0,12,7,
                1,15,13,8,10,3,7,4,12,5,6,11,0,14,9,2,
                7,11,4,1,9,12,14,2,0,6,10,13,15,3,5,8,
                2,1,14,7,4,10,8,13,15,12,9,0,3,5,6,11
            ]
        ]

        # P permutation
        p_table = [16,7,20,21,29,12,28,17,1,15,23,26,5,18,31,10,2,8,24,14,32,27,3,9,19,13,30,6,22,11,4,25]

        # Expand r
        r_expanded = self._permute(e_table, r)
        # XOR with subkey
        xr = [a ^ b for a, b in zip(r_expanded, subkey)]
        # Split into 8 groups of 6
        out_bits = []
        for i in range(8):
            chunk = xr[i*6:(i+1)*6]
            row = (chunk[0] << 1) | chunk[5]
            col = (chunk[1] << 3) | (chunk[2] << 2) | (chunk[3] << 1) | chunk[4]
            s_val = s_boxes[i][row*16 + col]
            for bit in range(4)[::-1]:
                out_bits.append((s_val >> bit) & 1)
        # P permutation
        p_out = self._permute(p_table, out_bits)
        return p_out

    def encrypt(self, data: bytes) -> bytes:
        """對單一 8-byte 區塊進行 DES 加密(ECB, 單區塊).

        例外:
        - TypeError: 非 bytes 輸入
        - ValueError: 長度非 8
        回傳: 加密後的 8-byte bytes
        """
        if not isinstance(data, (bytes, bytearray)):
            raise TypeError('data must be bytes')
        if len(data) != 8:
            raise ValueError('DES encrypt expects 8-byte block')

        # Initial permutation
        ip = [58,50,42,34,26,18,10,2,60,52,44,36,28,20,12,4,62,54,46,38,30,22,14,6,64,56,48,40,32,24,16,8,57,49,41,33,25,17,9,1,59,51,43,35,27,19,11,3,61,53,45,37,29,21,13,5,63,55,47,39,31,23,15,7]
        fp = [40,8,48,16,56,24,64,32,39,7,47,15,55,23,63,31,38,6,46,14,54,22,62,30,37,5,45,13,53,21,61,29,36,4,44,12,52,20,60,28,35,3,43,11,51,19,59,27,34,2,42,10,50,18,58,26,33,1,41,9,49,17,57,25]

        bits = self._bytes_to_bits(data)
        permuted = self._permute(ip, bits)
        l = permuted[:32]
        r = permuted[32:]

        for i in range(16):
            sub = self.subkeys[i]
            f_out = self._feistel(r, sub)
            new_r = [a ^ b for a, b in zip(l, f_out)]
            l = r
            r = new_r

        preoutput = r + l
        final_bits = self._permute(fp, preoutput)
        return self._bits_to_bytes(final_bits)


class DES:
    MODE_ECB = 1

    @staticmethod
    def new(key, mode=None):
        """工廠函式,回傳 _DES 實例以相容舊有介面.

        - 若傳入為 str,使用 latin-1 編碼轉為 bytes.
        - mode 目前僅為相容參數,未使用.
        """
        if isinstance(key, str):
            key = key.encode('latin-1')
        return _DES(key)

# 預設值(可用環境變數覆寫)
# 將可執行檔統一放到專案內的 `exe` 資料夾(相對於此檔案),使用環境變數可覆寫
BASE_DIR = Path(__file__).resolve().parent
EXE_DIR = BASE_DIR / 'exe'
# rustdesk 可執行檔路徑(相對或絕對)
RUSTDESK_APP = os.getenv('RUSTDESK_APP', str(EXE_DIR / 'rustdesk.exe'))
# 用於產生 RustDesk2.toml 的 rendezvous server 與 key(固定參數)
RUSTDESK_HOST = 'everdura.ddnsfree.com'
RUSTDESK_KEY = 'kCC8dq5x8uvEI+fpbIsTpYhCMaMbAxpYmGv6XtR7NsY='

# 是否在寫入 peers/{id}.toml 後把檔案設為唯讀(避免 RustDesk 立即覆寫)
RUSTDESK_SET_PEER_READONLY = False

# AnyDesk / TightVNC 可執行檔路徑
ANYDESK_APP = os.getenv('ANYDESK_APP', str(EXE_DIR / 'AnyDesk.exe'))
TIGHTVNC_APP = os.getenv('TIGHTVNC_APP', str(EXE_DIR / 'TightVNC.exe'))

# Base dir for resources when bundled with PyInstaller
VNC_BASE_DIR = getattr(sys, '_MEIPASS', None) or str(Path(__file__).resolve().parent)

def resource_path(filename: str) -> str:
    """取得打包後或開發模式下的資源檔案絕對路徑.

    參數:
    - filename: 相對於資源根目錄的檔案名稱

    回傳: 平台相容的絕對路徑字串
    """
    return os.path.join(VNC_BASE_DIR, filename)


def get_app_path(filename: str) -> str:
    """回傳應用程式相對的檔案路徑:
    - 若為 PyInstaller onefile/frozen,使用可執行檔所在資料夾 (sys.executable)
    - 否則使用原始 `BASE_DIR`(原始原始碼所在資料夾)
    """
    try:
        if getattr(sys, 'frozen', False):
            return os.path.join(os.path.dirname(sys.executable), filename)
    except Exception:
        pass
    return os.path.join(str(BASE_DIR), filename)


# Password paste helper removed per request: automated paste/Enter actions
# The repository may still contain helper binaries/scripts, but this application
# no longer invokes any automated password pasting.


def _find_excel_exe() -> str | None:
    """嘗試從登錄檢索 excel.exe 的路徑, 找不到則回傳 None."""
    try:
        # 優先查詢 HKLM / HKCU App Paths
        for hive in (winreg.HKEY_LOCAL_MACHINE, winreg.HKEY_CURRENT_USER):
            try:
                key = winreg.OpenKey(hive, r"SOFTWARE\Microsoft\Windows\CurrentVersion\App Paths\excel.exe")
                try:
                    exe_path, _ = winreg.QueryValueEx(key, None)
                    if exe_path:
                        return exe_path
                finally:
                    winreg.CloseKey(key)
            except FileNotFoundError:
                # 嘗試 WOW6432Node
                try:
                    key = winreg.OpenKey(hive, r"SOFTWARE\WOW6432Node\Microsoft\Windows\CurrentVersion\App Paths\excel.exe")
                    try:
                        exe_path, _ = winreg.QueryValueEx(key, None)
                        if exe_path:
                            return exe_path
                    finally:
                        winreg.CloseKey(key)
                except FileNotFoundError:
                    continue
        # 再試 HKEY_CLASSES_ROOT 的 ProgID
        try:
            key = winreg.OpenKey(winreg.HKEY_CLASSES_ROOT, r"Excel.Application\CurVer")
            winreg.CloseKey(key)
            # 若存在就回傳 None 表示安裝,但沒有取得路徑
            return None
        except Exception:
            return None
    except Exception:
        return None


def open_alldesk_excel(sheet_idx: int | None = None):
    """開啟 `Alldesk.xlsx`,若系統可用 COM automation,則嘗試選取指定工作表.

    參數:
    - sheet_idx: 1-based 的工作表索引,若為 None 則不指定.
    """
    xlsx = get_app_path('Alldesk.xlsx')
    if not os.path.exists(xlsx):
        log_and_show('找不到檔案', '找不到 Alldesk.xlsx', level='warning')
        return

    # delegate to helper that handles COM / exe / startfile
    try:
        open_excel_for_user(xlsx, sheet_idx=sheet_idx)
    except Exception:
        log_and_show('未安裝 Excel', '此電腦未偵測到 Microsoft Excel,無法以 Excel 開啟 Alldesk.xlsx', level='warning')


def open_excel_for_user(xlsx_path: str, sheet_idx: int | None = None, sheet_name: str | None = None):
    """開啟 Excel 檔供使用者編輯：

    行為：
    - 若可用 pywin32(com)，以 COM 開啟並嘗試 activate 指定 sheet (index 或 name)
    - 否則嘗試使用系統找到的 excel.exe 啟動或 `os.startfile`
    """
    exe_path = _find_excel_exe()

    # 優先使用 COM automation 控制 Excel(若可用)
    try:
        if win32com_client:
            try:
                excel = win32com_client.GetActiveObject('Excel.Application')
            except Exception:
                excel = win32com_client.Dispatch('Excel.Application')
            excel.Visible = True

            # 檢查檔案是否已開啟
            wb = None
            try:
                for w in excel.Workbooks:
                    try:
                        if os.path.normcase(w.FullName) == os.path.normcase(xlsx_path):
                            wb = w
                            break
                    except Exception:
                        continue
            except Exception:
                pass

            if wb is None:
                wb = excel.Workbooks.Open(xlsx_path)

            # 以 index 或 name 嘗試 activate
            if sheet_idx:
                try:
                    ws = wb.Worksheets(sheet_idx)
                    ws.Activate()
                except Exception:
                    pass
            elif sheet_name:
                try:
                    # Worksheets expects 1-based index or name
                    ws = wb.Worksheets(sheet_name)
                    ws.Activate()
                except Exception:
                    pass
            return
    except Exception:
        # 無法使用 COM，退到啟動檔案
        pass

    # 若有可執行檔路徑就用它開啟,否則用系統預設關聯
    if exe_path:
        try:
            subprocess.Popen([exe_path, xlsx_path])
            return
        except Exception:
            pass

    try:
        os.startfile(xlsx_path)
    except Exception:
        raise


def get_writable_dir() -> str:
    """回傳一個在此環境中可寫入的目錄.

    - 若為封裝後的執行檔(frozen),使用系統暫存目錄.
    - 開發模式則回傳此原始檔所在資料夾.
    """
    if getattr(sys, 'frozen', False):
        return tempfile.gettempdir()
    return os.path.dirname(__file__)


def encrypt_tightvnc_password(password: str) -> str:
    """將 TightVNC 的純文字密碼轉為 vnc 設定檔所使用的加密十六進位字串.

    演算法說明:
    - 取前 8 個 ASCII 字元,不足以 NUL 填充.
    - 使用 TightVNC 固定的 challenge bytes,對每個 byte 做 bit-reverse,
      將結果當作 DES key,使用 ECB 加密密碼區塊後回傳 hex 表示.
    """
    # take up to first 8 ASCII chars, pad with NULs
    pw = (password or '')[:8].encode('ascii', errors='ignore')
    pw = pw.ljust(8, b'\x00')

    # TightVNC uses a fixed challenge bytes; reverse bits in each challenge byte to form DES key
    challenge = [23, 82, 107, 6, 35, 78, 88, 7]

    def rev_bits_byte(b):
        b = ((b & 0xF0) >> 4) | ((b & 0x0F) << 4)
        b = ((b & 0xCC) >> 2) | ((b & 0x33) << 2)
        b = ((b & 0xAA) >> 1) | ((b & 0x55) << 1)
        return b

    key = bytes([rev_bits_byte(b) for b in challenge])
    cipher = DES.new(key, DES.MODE_ECB)
    return cipher.encrypt(pw).hex()


def create_header_row(parent, on_connect, with_port=False, default_port='5900'):
    """建立各遠端分頁共用的標頭區塊(輸入欄位與連接按鈕).

    參數:
    - parent: 要放置 header 的父容器 (tk widget)
    - on_connect: 當使用者按下「連接」按鈕時的回呼,會傳入 (id, pwd, port)
    - with_port: 是否顯示埠號輸入欄
    - default_port: 埠號欄位的預設值

    回傳: tuple (ent_id, ent_pwd, ent_port) - 若 `with_port` 為 False,則
    ent_port 會是 None.
    """
    header = ttk.Frame(parent)
    header.grid(row=0, column=0, columnspan=10, sticky='w')

    # 連接 ID
    f_id = ttk.Frame(header)
    f_id.pack(side='left', padx=10)
    tk.Label(f_id, text="連接ID:").pack(side='left')
    ent_id = tk.Entry(f_id, width=28)
    ent_id.pack(side='left', padx=6)

    # 密碼
    f_pwd = ttk.Frame(header)
    f_pwd.pack(side='left', padx=10)
    tk.Label(f_pwd, text="密碼:").pack(side='left')
    ent_pwd = tk.Entry(f_pwd, show='*', width=30)
    ent_pwd.pack(side='left', padx=6)

    # 連接按鈕
    # ent_port 可能尚未定義,故預先建立為 None,lambda 內再讀取
    ent_port = None
    def _on_click():
        on_connect(ent_id.get(), ent_pwd.get(), ent_port.get() if with_port and ent_port is not None else None)

    btn = tk.Button(header, text="連接", command=_on_click)
    btn.pack(side='left', padx=6)

    # 埠(可選)
    if with_port:
        f_port = ttk.Frame(header)
        f_port.pack(side='left', padx=10)
        tk.Label(f_port, text="埠:").pack(side='left')
        ent_port = tk.Entry(f_port, width=8)
        ent_port.pack(side='left', padx=6)
        ent_port.insert(0, default_port)

    return ent_id, ent_pwd, ent_port


def _sanitize_tag(s: str) -> str:
    """簡單清理 tag 字串，避免將程式碼片段或非常態文字顯示於 UI 上。"""
    if not isinstance(s, str):
        return ''
    v = s.strip()
    if not v:
        return ''
    low = v.lower()
    suspicious = ('import ', 'def ', 'class ', 'shutil', 'tkinter', 'pyinstaller', 'from ', 'subprocess')
    if any(tok in low for tok in suspicious):
        return ''
    if len(v) > 128:
        return ''
    non_print = sum(1 for ch in v if not ch.isprintable())
    if non_print > max(1, len(v) // 10):
        return ''
    return v



def _build_unilink_for_id(target_id: str, password: str | None = None) -> str:
    """建立 uni-link 字串，格式：rustdesk://connect/<id>?password=<pwd>（password 可為 None）"""
    try:
        tid = quote(str(target_id))
        if password:
            params = f"password={quote(str(password), safe='')}"
            return f"rustdesk://connect/{tid}?{params}"
        return f"rustdesk://connect/{tid}"
    except Exception:
        return f"rustdesk://connect/{target_id}"


def _find_flutter_runner_window(timeout: float = 3.0):
    user32 = ctypes.windll.user32
    class_name = "FLUTTER_RUNNER_WIN32_WINDOW"
    wnd_name = "RustDesk"
    start = time.time()
    while time.time() - start < timeout:
        try:
            hwnd = user32.FindWindowW(class_name, wnd_name)
            if hwnd and hwnd != 0:
                return hwnd
        except Exception:
            pass
        time.sleep(0.12)
    return None


def _send_unilink_to_flutter_runner(uni_link: str, timeout_ms: int = 2000) -> bool:
    """將 uni_link 用 WM_COPYDATA 發送到 Flutter runner（使用 SendMessageTimeoutW）。"""
    user32 = ctypes.windll.user32
    SMTO_ABORTIFHUNG = 0x0002
    WM_COPYDATA = 0x004A
    WM_USER = 0x0400

    hwnd = _find_flutter_runner_window(timeout=2.0)
    if not hwnd:
        return False

    class COPYDATASTRUCT(ctypes.Structure):
        _fields_ = [("dwData", ctypes.c_size_t), ("cbData", ctypes.c_ulong), ("lpData", ctypes.c_void_p)]

    try:
        # Use UTF-16LE wide string as core typically expects wide chars
        data_utf16 = (uni_link + '\x00').encode('utf-16le')
    except Exception:
        data_utf16 = (uni_link + '\x00').encode('utf-16le', errors='replace')

    buf = ctypes.create_string_buffer(data_utf16)
    cds = COPYDATASTRUCT()
    cds.dwData = WM_USER + 2
    cds.cbData = len(data_utf16)
    cds.lpData = ctypes.cast(buf, ctypes.c_void_p)

    try:
            if user32.SendMessageTimeoutW(hwnd, WM_COPYDATA, 0, ctypes.byref(cds), SMTO_ABORTIFHUNG, int(timeout_ms), None):
                return True
    except Exception:
        return False


def _find_window_for_id(target_id: str, timeout: float = 6.0):
    """嘗試找到視窗標題包含 target_id 的 RustDesk 視窗, 找到回傳 hwnd, 否則 None。"""
    user32 = ctypes.windll.user32
    WNDENUMPROC = ctypes.WINFUNCTYPE(ctypes.c_bool, ctypes.c_void_p, ctypes.c_void_p)
    buf = ctypes.create_unicode_buffer(1024)

    def enum_proc(hwnd, lParam):
        try:
            if user32.IsWindowVisible(hwnd):
                user32.GetWindowTextW(hwnd, buf, 1024)
                title = buf.value or ''
                if str(target_id) in title and 'RustDesk' in title:
                    found.append(hwnd)
                    return False
        except Exception:
            pass
        return True

    start = time.time()
    found = []
    while time.time() - start < timeout:
        found.clear()
        try:
            user32.EnumWindows(WNDENUMPROC(enum_proc), 0)
        except Exception:
            pass
        if found:
            return found[0]
        time.sleep(0.12)
    return None


def _find_password_dialog(timeout: float = 6.0):
    """尋找 RustDesk 的密碼提示對話視窗（以標題中含「密碼」或「需要密碼」為準）。
    回傳 hwnd 或 None。
    """
    user32 = ctypes.windll.user32
    WNDENUMPROC = ctypes.WINFUNCTYPE(ctypes.c_bool, ctypes.c_void_p, ctypes.c_void_p)
    buf = ctypes.create_unicode_buffer(1024)

    def enum_proc(hwnd, lParam):
        try:
            if not user32.IsWindowVisible(hwnd):
                return True
            user32.GetWindowTextW(hwnd, buf, 1024)
            title = (buf.value or '').strip()
            if not title:
                return True
            low = title.lower()
            if '密碼' in low or '需要密碼' in low or 'rustdesk 密碼' in low:
                found.append(hwnd)
                return False
        except Exception:
            pass
        return True

    start = time.time()
    found = []
    while time.time() - start < timeout:
        found.clear()
        try:
            user32.EnumWindows(WNDENUMPROC(enum_proc), 0)
        except Exception:
            pass
        if found:
            return found[0]
        time.sleep(0.12)
    return None


def _send_unilink_via_copydata(hwnd_target, uni_link: str) -> bool:
    """使用 WM_COPYDATA 將 uni_link 傳給指定視窗 (dwData = WM_USER+2)。"""
    user32 = ctypes.windll.user32
    WM_COPYDATA = 0x004A
    WM_USER = 0x0400

    class COPYDATASTRUCT(ctypes.Structure):
        _fields_ = [("dwData", ctypes.c_size_t), ("cbData", ctypes.c_ulong), ("lpData", ctypes.c_void_p)]

    def _try_send(data_bytes, encoding_name):
        try:
            buf = ctypes.create_string_buffer(data_bytes)
            cds = COPYDATASTRUCT()
            cds.dwData = WM_USER + 2
            cds.cbData = len(data_bytes)
            cds.lpData = ctypes.cast(buf, ctypes.c_void_p)
            res = user32.SendMessageW(hwnd_target, WM_COPYDATA, 0, ctypes.byref(cds))
            return int(res) != 0
        except Exception as e:
            return False

    try:
        data_utf8 = uni_link.encode('utf-8')
    except Exception:
        data_utf8 = uni_link.encode('utf-8', errors='replace')

    if _try_send(data_utf8, 'utf-8'):
        return True

    try:
        data_utf16 = (uni_link + '\x00').encode('utf-16le')
    except Exception:
        data_utf16 = (uni_link + '\x00').encode('utf-16le', errors='replace')

    if _try_send(data_utf16, 'utf-16le'):
        return True

    return False


def _set_clipboard_text(text: str) -> bool:
    """將 Unicode 文字放到系統剪貼簿（使用 WinAPI）。"""
    try:
        CF_UNICODETEXT = 13
        GMEM_MOVEABLE = 0x0002
        kernel32 = ctypes.windll.kernel32
        user32 = ctypes.windll.user32

        data = (text + '\x00').encode('utf-16le')
        hglobal = kernel32.GlobalAlloc(GMEM_MOVEABLE, len(data))
        if not hglobal:
            return False
        ptr = kernel32.GlobalLock(hglobal)
        if not ptr:
            kernel32.GlobalFree(hglobal)
            return False
        ctypes.memmove(ptr, data, len(data))
        kernel32.GlobalUnlock(hglobal)

        if not user32.OpenClipboard(None):
            kernel32.GlobalFree(hglobal)
            return False
        try:
            user32.EmptyClipboard()
            user32.SetClipboardData(CF_UNICODETEXT, hglobal)
        finally:
            user32.CloseClipboard()
        return True
    except Exception:
        return False


def _paste_via_keyboard_and_enter() -> bool:
    """模擬 Ctrl+V 然後 Enter。回傳 True 表示已發送鍵盤事件。"""
    try:
        user32 = ctypes.windll.user32
        KEYEVENTF_KEYUP = 0x0002
        VK_CONTROL = 0x11
        VK_V = 0x56
        VK_RETURN = 0x0D

        user32.keybd_event(VK_CONTROL, 0, 0, 0)
        user32.keybd_event(VK_V, 0, 0, 0)
        user32.keybd_event(VK_V, 0, KEYEVENTF_KEYUP, 0)
        user32.keybd_event(VK_CONTROL, 0, KEYEVENTF_KEYUP, 0)

        time.sleep(0.06)
        user32.keybd_event(VK_RETURN, 0, 0, 0)
        user32.keybd_event(VK_RETURN, 0, KEYEVENTF_KEYUP, 0)
        return True
    except Exception:
        return False


def _try_uia_set_password(hwnd, password: str) -> bool:
    """嘗試使用 pywinauto(UIA) 透過 hwnd 連到應用並設定密碼欄位的值。"""
    try:
        if PywinautoApplication is None:
            return False
        app = PywinautoApplication(backend='uia').connect(handle=hwnd)
        dlg = app.window(handle=hwnd)

        def _escape_for_type_keys(s: str) -> str:
            special = set('^%+~{}()[]')
            out = []
            for ch in s:
                if ch in special:
                    out.append('{' + ch + '}')
                else:
                    out.append(ch)
            return ''.join(out)

        escaped = _escape_for_type_keys(password)

        try:
            pw_edit = dlg.child_window(control_type='Edit')
            pw_edit.set_focus()
            pw_edit.type_keys(escaped, with_spaces=True, set_foreground=True)
            pw_edit.type_keys('{ENTER}')
            return True
        except Exception:
            try:
                edits = dlg.descendants(control_type='Edit')
                if edits:
                    edits[0].set_focus()
                    edits[0].type_keys(escaped, with_spaces=True, set_foreground=True)
                    edits[0].type_keys('{ENTER}')
                    return True
            except Exception:
                try:
                    if _set_clipboard_text(password):
                        try:
                            _force_foreground(hwnd)
                        except Exception:
                            pass
                        time.sleep(0.12)
                        if _paste_via_keyboard_and_enter():
                            return True
                except Exception:
                    return False
    except Exception:
        return False


def _force_foreground(hwnd: int) -> bool:
    """嘗試可靠地把指定 hwnd 帶到前景。

    先用 ShowWindow/SetForegroundWindow/BringWindowToTop，若失敗再嘗試 AttachThreadInput 打斷前景线程綁定。
    """
    try:
        user32 = ctypes.windll.user32
        SW_RESTORE = 9
        try:
            # 只在視窗為最小化時還原；避免對非最小化視窗呼叫 ShowWindow 改變其大小/狀態
            try:
                if user32.IsIconic(hwnd):
                    user32.ShowWindow(hwnd, SW_RESTORE)
            except Exception:
                pass
        except Exception:
            pass
        try:
            user32.SetForegroundWindow(hwnd)
            user32.BringWindowToTop(hwnd)
            return True
        except Exception:
            pass

        try:
            GetWindowThreadProcessId = user32.GetWindowThreadProcessId
            GetWindowThreadProcessId.restype = ctypes.c_ulong
            pid = ctypes.c_ulong()
            fg = user32.GetForegroundWindow()
            cur_tid = GetWindowThreadProcessId(fg, ctypes.byref(pid))
            tgt_pid = ctypes.c_ulong()
            tgt_tid = GetWindowThreadProcessId(hwnd, ctypes.byref(tgt_pid))

            AttachThreadInput = user32.AttachThreadInput
            AttachThreadInput.argtypes = [ctypes.c_ulong, ctypes.c_ulong, ctypes.c_bool]
            AttachThreadInput.restype = ctypes.c_bool

            # attach, set foreground, detach
            if AttachThreadInput(cur_tid, tgt_tid, True):
                try:
                    user32.SetForegroundWindow(hwnd)
                    user32.BringWindowToTop(hwnd)
                finally:
                    AttachThreadInput(cur_tid, tgt_tid, False)
                return True
        except Exception:
            pass
    except Exception:
        pass
    return False


def read_clients_from_sheet(sheet_names: list[str]) -> list[dict]:
    """從 `Alldesk.xlsx` 讀取指定命名工作表(大小寫不敏感)，回傳統一的 client dict 列表。

    回傳格式: [{"tag":..., "id":..., "pwd":..., "extra":{...}}, ...]
    `sheet_names` 可提供多個候選名稱(以優先順序匹配)。
    """
    excel_path = Path(get_app_path('Alldesk.xlsx'))
    clients = []
    if not excel_path.exists():
        return clients
    try:
        wb = open_workbook(str(excel_path))
        names = wb.sheetnames
        found = None
        for cand in sheet_names:
            found = next((s for s in names if str(s).strip().lower() == str(cand).strip().lower()), None)
            if found:
                break
        if not found:
            return clients
        ws = wb[found]
        rows = [tuple('' if v is None else v for v in r) for r in ws.iter_rows(values_only=True)]
        if not rows:
            return clients

        # 判斷是否 header row (檢查第一列是否包含常見欄名)
        first = [str(x).strip().lower() if x is not None else '' for x in rows[0]]
        header_tokens = ('id', 'item', 'name', 'tag', 'password', 'pwd', 'url', 'address', 'port')
        has_header = any(any(tok in cell for tok in header_tokens) for cell in first)

        if has_header:
            headers = [str(h).strip() if h is not None else '' for h in rows[0]]
            data_rows = rows[1:]
            for r in data_rows:
                # helper to find column by possible header names
                def get_by_headers(possibles):
                    for p in possibles:
                        for i, h in enumerate(headers):
                            if h.strip().lower() == p:
                                return '' if i >= len(r) or r[i] is None else str(r[i])
                    return None

                tag = get_by_headers(['item', 'tag', '設備名稱', 'name'])
                if tag is None:
                    tag = '' if len(r) < 1 or r[0] is None else str(r[0])
                id_ = get_by_headers(['id', 'url', 'address'])
                if id_ is None:
                    id_ = '' if len(r) < 2 or r[1] is None else str(r[1])
                pwd = get_by_headers(['password', 'pwd', 'pass', '密碼'])
                if pwd is None:
                    pwd = '' if len(r) < 3 or r[2] is None else str(r[2])

                extra = {}
                for i, h in enumerate(headers):
                    key = h if h else f'col{i+1}'
                    extra[key] = '' if i >= len(r) or r[i] is None else r[i]

                clients.append({'tag': str(tag), 'id': str(id_), 'pwd': str(pwd), 'extra': extra})
        else:
            for r in rows:
                tag = '' if len(r) < 1 or r[0] is None else str(r[0])
                id_ = '' if len(r) < 2 or r[1] is None else str(r[1])
                pwd = '' if len(r) < 3 or r[2] is None else str(r[2])
                extra = {}
                for i, v in enumerate(r[3:], start=4):
                    extra[f'col{i}'] = '' if v is None else v
                clients.append({'tag': str(tag), 'id': str(id_), 'pwd': str(pwd), 'extra': extra})
    except Exception:
        return []
    return clients


def open_workbook(path: str):
    """通用 workbook 打開：目前以 openpyxl 的 load_workbook 為主。

    若日後需加入 COM 讀取邏輯，可在此封裝選擇。
    回傳 workbook 或 raise exception。
    """
    return load_workbook(filename=str(path), read_only=True, data_only=True)


def normalize_client_fields(client: dict) -> dict:
    """統一處理 client dict 欄位：id 去除 '.0'、數字轉字串、空值處理。"""
    out = {'tag': '', 'id': '', 'pwd': '', 'extra': {}}
    if not isinstance(client, dict):
        return out
    try:
        tag = client.get('tag', '') or ''
        id_ = client.get('id', '') if client.get('id', '') is not None else ''
        pwd = client.get('pwd', '') or ''
        extra = client.get('extra', {}) or {}
    except Exception:
        return out
    try:
        if isinstance(id_, (int, float)):
            id_ = str(id_)
        id_ = str(id_).strip()
        if id_.endswith('.0'):
            id_ = id_[:-2]
    except Exception:
        id_ = ''
    try:
        tag = str(tag).strip()
    except Exception:
        tag = ''
    try:
        # 確保密碼去除前後空白，避免 Excel 格式問題導致比對失敗
        pwd = str(pwd).strip()
    except Exception:
        pwd = ''
    out['tag'] = tag
    out['id'] = id_
    out['pwd'] = pwd
    out['extra'] = extra
    return out


def get_extra_field(client: dict, candidates: list[str]):
    """從 client['extra'] 以不區分大小寫的 key 找出第一個匹配的值。"""
    try:
        extra = client.get('extra', {}) or {}
        for k, v in extra.items():
            if str(k).strip().lower() in candidates:
                return v
    except Exception:
        pass
    return ''


def launch_process(cmd, cwd=None, creationflags=None, timeout=None, stdout=None, stderr=None):
    """統一啟動外部程式的 helper，回傳 Popen 物件。

    - `cmd` 可為 list 或字串
    - 若 `creationflags` 未指定,預設 None
    - `timeout` 不會阻塞；僅在需要時可用 proc.wait(timeout)
    """
    try:
        proc = subprocess.Popen(cmd, cwd=cwd or get_writable_dir(), creationflags=creationflags, stdout=stdout, stderr=stderr)
        return proc
    except Exception as e:
        return None


def create_client_buttons(container, clients: list[dict], on_connect, cols: int = 10, btn_font=('微軟正黑體', 10)):
    """在 container 上建立按鈕網格，`on_connect` 會收到整個 client dict 作為參數。

    - 跳過 header-like rows
    - 使用 `_sanitize_tag` 清理顯示文字
    """
    btn_container = ttk.Frame(container)
    btn_container.grid(row=2, column=0, columnspan=10, sticky='w')
    row = 0
    col = 0
    for client in clients:
        client = normalize_client_fields(client)
        try:
            tag = client.get('tag', '') or ''
        except Exception:
            tag = ''
        try:
            client_id = client.get('id', '') or ''
        except Exception:
            client_id = ''
        try:
            pwd = client.get('pwd', '') or ''
        except Exception:
            pwd = ''

        tag = _sanitize_tag(tag)
        if isinstance(client_id, (int, float)):
            client_id = str(client_id)
        client_id = client_id.strip()
        if client_id.endswith('.0'):
            client_id = client_id[:-2]

        if isinstance(tag, str) and tag.strip().lower() in ('設備名稱', 'id', 'item', 'name'):
            continue
        if isinstance(client_id, str) and client_id.strip().lower() in ('設備名稱', 'id', 'item', 'name'):
            continue
        if not tag and not client_id:
            continue

        try:
            tk.Button(btn_container, text=f"{tag}\n{client_id}", font=btn_font, width=15, height=4,
                      command=(lambda c=client: on_connect(c))).grid(row=row, column=col, padx=3, pady=3)
        except Exception:
            pass
        col += 1
        if col >= cols:
            col = 0
            row += 1


def _atomic_write_text(path: str, data: str, encoding: str = 'utf-8') -> None:
    """以原子方式寫入文字檔:
    - 先在同一目錄建立暫存檔,寫入並 fsync
    - 以 os.replace 原子取代目標檔
    - 嘗試在 replace 前解除目標檔的唯讀屬性以避免在 Windows 上失敗
    """
    dirp = os.path.dirname(path) or get_writable_dir()
    fd = None
    tmp = None
    try:
        fd, tmp = tempfile.mkstemp(prefix=os.path.basename(path) + '.tmp.', dir=dirp)
        with os.fdopen(fd, 'w', encoding=encoding, newline='\n') as fw:
            fd = None
            fw.write(data)
            try:
                fw.flush()
                os.fsync(fw.fileno())
            except Exception:
                pass
        # 確保目標可寫(若存在且為唯讀),再 replace
        try:
            if os.path.exists(path):
                try:
                    os.chmod(path, stat.S_IWRITE)
                except Exception:
                    pass
        except Exception:
            pass
        os.replace(tmp, path)
        tmp = None
    finally:
        try:
            if fd is not None:
                try:
                    os.close(fd)
                except Exception:
                    pass
        except Exception:
            pass
        try:
            if tmp and os.path.exists(tmp):
                os.remove(tmp)
        except Exception:
            pass

class RustDesk():
    """RustDesk 分頁:從 Excel 載入 client 並發起 RustDesk 連線.

    精簡且安全的實作:
    - 將 per-peer 的 view_style 等設定寫入 `peers/{ID}.toml`(原子寫入).
    - 只把 rendezvous/relay/key 等全域設定寫入 `RustDesk2.toml`(原子寫入),不放 per-peer 密碼.
    - 連線時仍以 `--password` 參數傳入密碼(不在全域檔案放明文).
    """
    def __init__(self, notebook: ttk.Notebook):
        self.init_rustdesk(notebook)

    def init_rustdesk(self, notebook: ttk.Notebook):
        # 在 runtime 決定 rustdesk 可執行檔路徑（避免打包時 import-time 常數路徑失效）
        # 優先使用環境變數, 否則嘗試相對於執行檔或專案的 exe/rustdesk.exe
        try:
            app = os.getenv('RUSTDESK_APP')
            if not app:
                # get_app_path 會在 frozen 時回傳可執行檔所在目錄
                app = get_app_path(os.path.join('exe', 'rustdesk.exe'))
            # 若找不到, 在 frozen 模式嘗試可執行檔同目錄下的 rustdesk.exe
            if not os.path.exists(app):
                if getattr(sys, 'frozen', False):
                    maybe = os.path.join(os.path.dirname(sys.executable), 'rustdesk.exe')
                    if os.path.exists(maybe):
                        app = maybe
        except Exception:
            app = RUSTDESK_APP

        # 使用共用 Excel 讀取 helper
        clients = read_clients_from_sheet(['rustdesk'])
        self.exec_target = os.path.normpath(app)
        self.clients = clients
        self.frame = ttk.Frame(notebook)
        notebook.add(self.frame, text='RustDesk')

    def _prepare_rustdesk_conf(self, client_id: str, password: str):
        # stat imported at module top
        appdata = os.getenv('APPDATA')
        if not appdata:
            return
        cfg_dir = os.path.join(appdata, 'RustDesk', 'config')
        peers_dir = os.path.join(cfg_dir, 'peers')
        Path(peers_dir).mkdir(parents=True, exist_ok=True)

        # 1. 處理 ID(避免 Excel 轉成浮點並出現 .0)
        try:
            target_id = '' if client_id is None else str(client_id).strip()
        except Exception:
            target_id = ''
        if target_id.endswith('.0'):
            target_id = target_id[:-2]

        peer_file = os.path.join(peers_dir, f"{target_id}.toml")

        # 若 peer 設定已存在,記錄後跳過 peer 寫入的預設邏輯已改為內容比對

        # 2. 精準 peer.toml 內容(強制 view_style = 'adaptive')
        peer_content = (
            "password = []\n"
            "size = [\n"
            "    0,\n"
            "    0,\n"
            "    0,\n"
            "    0,\n"
            "]\n"
            "size_ft = [\n"
            "    0,\n"
            "    0,\n"
            "    0,\n"
            "    0,\n"
            "]\n"
            "size_pf = [\n"
            "    0,\n"
            "    0,\n"
            "    0,\n"
            "    0,\n"
            "]\n"
            "view_style = 'adaptive'\n"
            "scroll_style = 'scrollauto'\n"
            "edge_scroll_edge_thickness = 100\n"
            "image_quality = 'balanced'\n"
            "custom_image_quality = [50]\n"
            "show_remote_cursor = false\n"
            "lock_after_session_end = false\n"
            "terminal-persistent = false\n"
            "privacy_mode = false\n"
            "allow_swap_key = false\n"
            "port_forwards = []\n"
            "direct_failures = 0\n"
            "disable_audio = false\n"
            "disable_clipboard = false\n"
            "enable-file-copy-paste = true\n"
            "show_quality_monitor = false\n"
            "follow_remote_cursor = false\n"
            "follow_remote_window = false\n"
            "keyboard_mode = 'map'\n"
            "view_only = false\n"
            "show_my_cursor = false\n"
            "sync-init-clipboard = false\n"
            "trackpad-speed = 100\n\n"
            "[options]\n"
            "codec-preference = 'auto'\n"
            "swap-left-right-mouse = ''\n"
            "collapse_toolbar = ''\n"
            "custom-fps = '30'\n"
            "zoom-cursor = ''\n"
            "i444 = ''\n\n"
            "[ui_flutter]\n"
            "wm_RemoteDesktop = '{\"width\":1270.0,\"height\":710.0,\"offsetWidth\":1270.0,\"offsetHeight\":710.0,\"isMaximized\":true,\"isFullscreen\":false}'\n\n"
            "[info]\n"
            "username = 'VMM'\n"
            "hostname = 'soyal-pc'\n"
            "platform = 'Windows'\n\n"
            "[transfer]\n"
            "write_jobs = []\n"
            "read_jobs = []\n"
        )

        # 3. 處理 RustDesk2.toml(僅在真的沒設定時寫入,避免觸發重新載入)
        cfg_file = os.path.join(cfg_dir, 'RustDesk2.toml')
        need_write_cfg = True
        if os.path.exists(cfg_file):
            try:
                with open(cfg_file, 'r', encoding='utf-8') as fr:
                    if RUSTDESK_HOST in fr.read():
                        need_write_cfg = False
            except Exception:
                pass

        if need_write_cfg:
            try:
                cfg_data = (
                    f"rendezvous_server = '{RUSTDESK_HOST}:21116'\n"
                    "nat_type = 1\n"
                    "serial = 0\n"
                    "unlock_pin = ''\n"
                    "trusted_devices = ''\n\n"
                    "[options]\n"
                    f"relay-server = '{RUSTDESK_HOST}:21117'\n"
                    f"custom-rendezvous-server = '{RUSTDESK_HOST}:21116'\n"
                    "local-ip-addr = ''\n"
                    f"key = '{RUSTDESK_KEY}'\n"
                    "av1-test = 'Y'\n"
                )
                tmp_cfg = cfg_file + '.tmp'
                try:
                    with open(tmp_cfg, 'w', encoding='utf-8', newline='\n') as fw:
                        fw.write(cfg_data)
                        try:
                            fw.flush()
                            os.fsync(fw.fileno())
                        except Exception:
                            pass
                    os.replace(tmp_cfg, cfg_file)
                except Exception:
                    try:
                        if os.path.exists(cfg_file):
                            os.remove(cfg_file)
                        os.replace(tmp_cfg, cfg_file)
                    except Exception:
                        pass
            except Exception:
                pass

        # 4. 先讀取 peer 檔內容;若內容已正確,early return,絕對不要更動檔案
        try:
            if os.path.exists(peer_file):
                try:
                    with open(peer_file, 'r', encoding='utf-8') as fr:
                        current = fr.read()
                    if "view_style = 'adaptive'" in current :
                        return
                except Exception:
                    # 無法讀取時繼續到寫入流程
                    pass
        except Exception:
            pass

        # 5. 只有在檔案不存在或內容不符時才寫入：使用原子寫入避免競態與中間暫存檔
        try:
            _atomic_write_text(peer_file, peer_content, encoding='utf-8')
        except Exception as e:
            pass

    def run_rustdesk(self, client_id, password):
        exec_target = self.exec_target

        # 確認 exec_target 是否存在；若不存在，嘗試大小寫不敏感或常見位置搜尋
        try:
            if not exec_target or not os.path.exists(exec_target):
                candidates = []

                def _add_if_exists(p):
                    try:
                        if p and os.path.exists(p):
                            candidates.append(p)
                    except Exception:
                        pass

                # 明確的優先順序：先搜尋 _internal\exe，接著搜尋 exe，再搜尋其他常見位置
                ordered_roots = []
                try:
                    if getattr(sys, 'frozen', False):
                        exe_dir = os.path.dirname(sys.executable)
                        ordered_roots = [
                            os.path.join(exe_dir, '_internal', 'exe'),
                            os.path.join(exe_dir, 'exe'),
                            os.path.join(exe_dir, '_internal'),
                            exe_dir,
                        ]
                    else:
                        ordered_roots = [
                            os.path.join(str(BASE_DIR), '_internal', 'exe'),
                            os.path.join(str(BASE_DIR), 'exe'),
                            os.path.join(str(BASE_DIR), '_internal'),
                            str(BASE_DIR),
                        ]
                except Exception:
                    ordered_roots = [os.path.join(str(BASE_DIR), '_internal', 'exe'), os.path.join(str(BASE_DIR), 'exe'), str(BASE_DIR)]

                for root in ordered_roots:
                    try:
                        _add_if_exists(os.path.join(root, 'rustdesk.exe'))
                        _add_if_exists(os.path.join(root, 'RustDesk.exe'))
                        if os.path.isdir(root):
                            for fn in os.listdir(root):
                                if fn.lower() == 'rustdesk.exe':
                                    _add_if_exists(os.path.join(root, fn))
                    except Exception:
                        pass

                # 也檢查 resource_path / get_app_path 對應的路徑
                try:
                    _add_if_exists(resource_path(os.path.join('_internal', 'exe', 'rustdesk.exe')))
                    _add_if_exists(resource_path(os.path.join('exe', 'rustdesk.exe')))
                    _add_if_exists(get_app_path(os.path.join('_internal', 'exe', 'rustdesk.exe')))
                    _add_if_exists(get_app_path(os.path.join('exe', 'rustdesk.exe')))
                except Exception:
                    pass

                # 環境變數作為最後保底
                try:
                    _add_if_exists(os.getenv('RUSTDESK_APP'))
                except Exception:
                    pass

                if candidates:
                    exec_target = candidates[0]
                else:
                    pass
        except Exception:
            pass

        # prepare config
        self._prepare_rustdesk_conf(client_id, password)

        # 如果是在 onedir (frozen 且非 onefile) 或是執行檔位於專案的 exe 目錄，
        # 則先把執行檔複製到臨時目錄（隨機名稱）再啟動，以強制建立新實例，避免 single-instance 攔截。
        try:
            onefile_extracted = getattr(sys, '_MEIPASS', None) is not None
            should_copy = False
            if getattr(sys, 'frozen', False) and not onefile_extracted:
                should_copy = True
            else:
                try:
                    # 若 exec_target 在專案的 exe 資料夾下，也建議複製以達致與 onefile 一致行為
                    if exec_target and os.path.commonpath([os.path.abspath(exec_target), str(EXE_DIR)]) == os.path.abspath(str(EXE_DIR)):
                        should_copy = True
                except Exception:
                    pass

            if should_copy and exec_target and os.path.exists(exec_target):
                try:
                    tmp_name = f"rustdesk_{uuid.uuid4().hex}.exe"
                    tmp_path = os.path.join(tempfile.gettempdir(), tmp_name)
                    shutil.copy2(exec_target, tmp_path)
                    os.chmod(tmp_path, os.stat(tmp_path).st_mode | stat.S_IREAD | stat.S_IWRITE | stat.S_IEXEC)
                    exec_target = tmp_path
                except Exception:
                    # 若複製失敗，繼續使用原始 exec_target
                    pass
        except Exception:
            pass

        # 1) launch (start without passing password; use UIA/clipboard/uni-link to supply password)
        cmd = [exec_target, '--connect', str(client_id)]
        try:
            proc = launch_process(cmd, creationflags=subprocess.CREATE_NEW_CONSOLE)
            pass
        except Exception as e:
            proc = None
            pass

        # 2) wait for connection window
        hwnd = _find_window_for_id(str(client_id), timeout=6.0)
        # 記錄是否在最初檢查時就已找到視窗，之後若已找到就不再等待後續的 long-time 查找
        initial_found = bool(hwnd)

        # 若找到 connection window，嘗試最大化視窗以便使用者觀看
        try:
            if hwnd:
                try:
                    user32 = ctypes.windll.user32
                    SW_MAXIMIZE = 3
                    # 明確要求最大化（避免使用 SW_SHOW 以免改變非最大化狀態）
                    user32.ShowWindow(hwnd, SW_MAXIMIZE)
                    try:
                        user32.SetForegroundWindow(hwnd)
                        user32.BringWindowToTop(hwnd)
                    except Exception:
                        pass
                except Exception:
                    pass
        except Exception:
            pass

        if not hwnd:
            # fallback: launch without password (original behavior)
            try:
                launch_process([exec_target, '--connect', str(client_id)], creationflags=subprocess.CREATE_NEW_CONSOLE)
                pass
                return True
            except Exception as e:
                pass
                return False

        # 3) build uni-link and try Flutter runner
        uni = _build_unilink_for_id(client_id, password)
        try:
            if _send_unilink_to_flutter_runner(uni):
                # 若最初已找到視窗，直接復用該 HWND，避免再次等待 timeout
                hwnd2 = hwnd if initial_found else _find_window_for_id(str(client_id), timeout=6.0)
                if hwnd2:
                    # bring the specific connection window to front before attempting UIA/paste
                    try:
                        _force_foreground(hwnd2)
                        time.sleep(0.08)
                    except Exception:
                        pass
                    try:
                        if _try_uia_set_password(hwnd2, str(password)):
                            pass
                            return True
                    except Exception:
                        pass
                    # clipboard fallback for the runner window
                    try:
                        if _set_clipboard_text(str(password)):
                            time.sleep(0.12)
                            if _paste_via_keyboard_and_enter():
                                pass
                                return True
                    except Exception:
                        pass
        except Exception:
            pass

        # 4) try WM_COPYDATA to connection window
        try:
            if _send_unilink_via_copydata(hwnd, uni):
                # Prefer acting on the explicit password dialog if it appears.
                try:
                    pwd_hwnd = _find_password_dialog(timeout=2.0)
                except Exception:
                    pwd_hwnd = None
                if pwd_hwnd:
                    try:
                        _force_foreground(pwd_hwnd)
                    except Exception:
                        pass
                    try:
                        if _try_uia_set_password(pwd_hwnd, str(password)):
                            pass
                            return True
                    except Exception:
                        pass
                else:
                    # fallback: try UIA on connection window but avoid forcing generic main window to foreground
                    try:
                        if _try_uia_set_password(hwnd, str(password)):
                            pass
                            return True
                    except Exception:
                        pass
                try:
                    if _set_clipboard_text(str(password)):
                        time.sleep(0.12)
                        if _paste_via_keyboard_and_enter():
                            pass
                            return True
                except Exception:
                    pass
                return True
        except Exception:
            pass

        # 5) final fallback: start without password on CLI (preserve original behavior)
        # 如果最初已找到視窗 (initial_found)，表示我們曾成功看見連線視窗，
        # 使用者可能是手動關閉視窗；為避免自動重啟，在這種情況下不做 CLI fallback。
        if not (locals().get('initial_found') or False):
            try:
                launch_process([exec_target, '--connect', str(client_id)], creationflags=subprocess.CREATE_NEW_CONSOLE)
                pass
                return True
            except Exception:
                pass

        return True

    def set_elements_rustdesk(self):
        create_header_row(self.frame, on_connect=lambda cid, pwd, _: self.run_rustdesk(cid, pwd), with_port=False)
        create_client_buttons(self.frame, self.clients, on_connect=lambda c: self.run_rustdesk(c.get('id'), c.get('pwd')))

class AnyDesk():
    """AnyDesk 分頁:從 Excel 載入 client 並啟動 AnyDesk 連線.

    主要職責:
    - 從 `Alldesk.xlsx` 的 'anydesk' 工作表讀取客戶清單.
    - 在啟動 AnyDesk 前於 %AppData%/AnyDesk 寫入 `user.conf`,以控制視圖模式.
    """
    def __init__(self, notebook: ttk.Notebook):
        """建立 AnyDesk 分頁物件並初始化其 UI 與資料.

        傳入 `notebook` 並呼叫 `init_anydesk` 讀取 Excel 並準備按鈕與執行檔路徑.
        """
        self.init_anydesk(notebook)

    def init_anydesk(self, notebook: ttk.Notebook):
        """初始化 AnyDesk 分頁:

        - 讀取 `Alldesk.xlsx` 的 'anydesk' 工作表(或第二張表),
          解析成 (tag, id, password) 的 client 列表.
        - 正規化 AnyDesk 可執行檔路徑並建立 UI 容器.
        """
        app: str = ANYDESK_APP
        clients = read_clients_from_sheet(['anydesk'])
        exec_target = os.path.normpath(app)

        self.exec_target = exec_target
        self.clients = clients
        self.frame = ttk.Frame(notebook)
        notebook.add(self.frame, text = 'AnyDesk')

    def _prepare_anydesk_conf(self, client_id: str):
        r"""在 %AppData%\AnyDesk 下建立 `user.conf` 並設定 viewmode.

        只寫入最小內容:`ad.session.viewmode=<client_id>:2`,用以在啟動 AnyDesk 時
        影響視窗顯示模式(例如強制開啟為檢視模式或預設尺寸).
        """
        appdata = os.getenv('APPDATA')
        if not appdata:
            return
        anydesk_dir = os.path.join(appdata, 'AnyDesk')
        Path(anydesk_dir).mkdir(parents=True, exist_ok=True)
        conf_file = os.path.join(anydesk_dir, 'user.conf')
        try:
            with open(conf_file, 'w', encoding='utf-8') as fw:
                fw.write(f"ad.session.viewmode={client_id}:2\n")
                fw.write("ad.installation.reminder_enabled=false\n")
                fw.write("ad.ui.inst_info_count=100\n")
                fw.write("ad.ui.last_reminder_time=1768860673\n")
                fw.write("ad.ui.install_skipped=true\n")
                fw.write("ad.features.install=false\n")
        except Exception:
            pass

    def run_anydesk(self, client_id, password):
        r"""啟動 AnyDesk 連線(AnyDesk 專用).

        步驟:
        1. 呼叫 `_prepare_anydesk_conf`,將 viewmode 寫入 `%APPDATA%\AnyDesk\user.conf`.
        2. 以非同步方式呼叫 AnyDesk,並透過命令列管道傳入密碼.
        """
        exec_target = self.exec_target
        # 在啟動 AnyDesk 前,先寫入 user.conf 以設定 viewmode
        self._prepare_anydesk_conf(client_id)

        # 使用 cmd 管道傳入密碼並以非同步方式啟動 AnyDesk
        try:
            # 優先嘗試以系統管理員權限啟動 (會顯示 UAC 提示)
            # 使用 elevated cmd 來傳入密碼管道，若失敗則回退到原先的 subprocess 行為
            if client_id:
                params = f'/c echo {password} | "{exec_target}" "{client_id}" --with-password'
            else:
                params = f'/c echo {password} | "{exec_target}" --with-password'
            try:
                ctypes.windll.shell32.ShellExecuteW(None, 'runas', 'cmd.exe', params, None, 0)
            except Exception:
                # 若 ShellExecuteW 無法呼叫 (或使用者取消 UAC)，回退原本的非同步啟動
                command = f'cmd /c echo {password} | "{exec_target}" "{client_id}" --with-password'
                subprocess.Popen(command, creationflags = subprocess.CREATE_NO_WINDOW)
        except Exception:
            try:
                # fallback: start without piping
                cmd = [exec_target, str(client_id)] if client_id else [exec_target]
                subprocess.Popen(cmd, creationflags=subprocess.CREATE_NEW_CONSOLE)
            except Exception as e:
                pass

    def set_elements_anydesk(self):
        """建立 AnyDesk 分頁的 UI(header + 按鈕)."""
        create_header_row(self.frame, on_connect=lambda cid, pwd, _: self.run_anydesk(cid, pwd), with_port=False)
        create_client_buttons(self.frame, self.clients, on_connect=lambda c: self.run_anydesk(c.get('id'), c.get('pwd')))

class TightVNC():
    """TightVNC 分頁:從 Alldesk.xlsx 第3張工作表載入項目並啟動 VNC 連線.

    欄位對應:
    - Item: 顯示在按鈕上的名稱
    - URL: 目標主機(按鈕上顯示)
    - Password: 連線密碼(按鈕上不顯示)
    - Port: 連接埠(按鈕上不顯示)
    """
    def __init__(self, notebook: ttk.Notebook):
        app = 'vnc'
        # 使用共用 helper 嘗試讀取 'vnc' 或 'tightvnc' 工作表
        clients = read_clients_from_sheet(['vnc', 'tightvnc'])

        self.exec_target = TIGHTVNC_APP
        self.clients = clients
        self.frame = ttk.Frame(notebook)
        notebook.add(self.frame, text = 'TightVNC')

    def _prepare_and_launch_tightvnc(self, host, port, password):
        r"""讀取 vnc.vnc 範本,替換 connection/ options,寫出並啟動 TightVNC."""
        vnc_source = resource_path('vnc.vnc')
        if os.path.exists(vnc_source):
            try:
                with open(vnc_source, 'r', encoding='utf-8') as f:
                    lines = f.readlines()
            except Exception:
                lines = []
        else:
            lines = []

        out = []
        in_conn = False
        replaced = {'host': False, 'port': False, 'password': False}
        for i, line in enumerate(lines):
            s = line.strip()
            if s.lower() == '[connection]':
                in_conn = True
                out.append(line)
                continue
            if in_conn:
                if s.startswith('[') and s.endswith(']'):
                    in_conn = False
                    out.append(line)
                    continue
                if s.lower().startswith('host='):
                    out.append(f'host={host}\n')
                    replaced['host'] = True
                    continue
                if s.lower().startswith('port='):
                    out.append(f'port={port}\n')
                    replaced['port'] = True
                    continue
                if s.lower().startswith('password='):
                    if password:
                        enc_pw = encrypt_tightvnc_password(password)
                        out.append(f'password={enc_pw}\n')
                        replaced['password'] = True
                    else:
                        out.append(line)
                    continue
            out.append(line)

        if not any(l.strip().lower() == '[connection]' for l in out):
            conn_block = ["[connection]\n", f"host={host}\n", f"port={port}\n"]
            if password:
                enc_pw = encrypt_tightvnc_password(password)
                conn_block.append(f"password={enc_pw}\n")
            out = conn_block + ['\n'] + out
        else:
            if not (replaced['host'] and replaced['port'] and replaced['password']):
                new_out = []
                i = 0
                while i < len(out):
                    new_out.append(out[i])
                    if out[i].strip().lower() == '[connection]':
                        j = i + 1
                        consume = []
                        while j < len(out) and not out[j].strip().startswith('['):
                            consume.append(out[j])
                            j += 1
                        conn_lines = [f'host={host}\n', f'port={port}\n']
                        if password:
                            enc_pw = encrypt_tightvnc_password(password)
                            conn_lines.append(f'password={enc_pw}\n')
                        else:
                            for c in consume:
                                if c.strip().lower().startswith('password='):
                                    conn_lines.append(c)
                                    break
                        new_out.extend(conn_lines)
                        i = j
                        continue
                    i += 1
                out = new_out

                def ensure_options(lines):
                    has_options = False
                    i = 0
                    while i < len(lines):
                        if lines[i].strip().lower() == '[options]':
                            has_options = True
                            j = i + 1
                            opts = {}
                            while j < len(lines) and not lines[j].strip().startswith('['):
                                s = lines[j].strip()
                                if '=' in s:
                                    k, v = s.split('=', 1)
                                    opts[k.strip().lower()] = v.strip()
                                j += 1
                            opts['viewonly'] = '0'
                            opts['shared'] = '1'
                            opts['swapmouse'] = opts.get('swapmouse', '0')
                            new_block = ['[options]\n']
                            for k, v in opts.items():
                                new_block.append(f'{k}={v}\n')
                            lines[i:j] = new_block
                            break
                        i += 1
                    if not has_options:
                        opts_block = ['[options]\n', 'viewonly=0\n', 'shared=1\n', 'swapmouse=0\n', '\n']
                        lines.extend(opts_block)
                    return lines

                out = ensure_options(out)

        try:
            Path(EXE_DIR).mkdir(parents=True, exist_ok=True)
        except Exception:
            pass
        out_path = os.path.join(str(EXE_DIR), 'vnc.vnc')
        try:
            with open(out_path, 'w', encoding='utf-8') as f:
                f.writelines(out)
        except Exception:
            return

        exe_path = resource_path('TightVNC.exe')
        if not os.path.exists(exe_path):
            exe_path = TIGHTVNC_APP
        if not os.path.exists(exe_path):
            exe_path = 'TightVNC.exe'
        args = [exe_path, f'-optionsfile={out_path}', '-showcontrols=no']
        try:
            subprocess.Popen(args, cwd=get_writable_dir())
        except Exception:
            pass

    def run_tightvnc(self, item, url, password, port):
        """啟動 TightVNC 連線的高階介面."""
        host = url or ''
        prt = port or '5900'
        self._prepare_and_launch_tightvnc(host, prt, password)

    def set_elements_tightvnc(self):
        create_header_row(
            self.frame,
            on_connect=lambda cid, pwd, port: self.run_tightvnc('', cid, pwd, port),
            with_port=True,
            default_port='5900'
        )
        # 使用共用 buttons helper；on_connect 會得到整個 client dict
        create_client_buttons(self.frame, self.clients,
                      on_connect=lambda c: self.run_tightvnc(c.get('tag'), c.get('id'), c.get('pwd'), str(get_extra_field(c, ['port', '埠', '埠號']) or '')),
                      cols=10)
    

gui = tk.Tk()
gui.title('Alldesk')

# 調整 Notebook 標籤字型:加大並改為粗體以便與 UI 一致
style = ttk.Style()
# 為了讓 tab 的背景/前景 mapping 生效,嘗試使用 'clam' 主題(較支援 element 顏色客製化)
try:
    if 'clam' in style.theme_names():
        style.theme_use('clam')
except Exception:
    pass
tab_font = tkfont.Font(family='微軟正黑體', size=11, weight='bold')
style.configure('Big.TNotebook.Tab', font=tab_font, padding=[12, 6], background='#f0f0f0', foreground='black')
# 確保 Notebook 本體與 tab 的預設背景一致
try:
    style.configure('TNotebook', background='#f0f0f0')
    style.configure('TNotebook.Tab', background='#f0f0f0')
except Exception:
    pass
# 當 tab 被選取時顯示黑底白字;未選取則為淺灰底黑字
style.map('Big.TNotebook.Tab',
    background=[('selected', 'black'), ('!selected', '#f0f0f0')],
    foreground=[('selected', 'white'), ('!selected', 'black')]
)

# 使用一個容器,將 `Notebook` 放左邊,右邊放一個 `EXCEL` 按鈕
container = ttk.Frame(gui)
container.pack(fill='both', expand=True)
notebook = ttk.Notebook(container, style='Big.TNotebook')
notebook.pack(side='left', fill='both', expand=True)

rustdesk = RustDesk(notebook)
rustdesk.set_elements_rustdesk()

anydesk = AnyDesk(notebook)
anydesk.set_elements_anydesk()

tightvnc = TightVNC(notebook)
tightvnc.set_elements_tightvnc()

# 新增一個與其他分頁相同風格的 `EXCEL` 分頁(放在最右側)
excel_frame = ttk.Frame(notebook)
notebook.add(excel_frame, text='EXCEL')

# 追蹤上一個選取的 tab id,點選 EXCEL 分頁時開啟檔案並還原為上一分頁
_last_tab = {'id': notebook.select()}

def _on_tab_changed(event):
    """Notebook tab 變更事件處理.

    若使用者切換到 'EXCEL' 分頁,則開啟 `Alldesk.xlsx`(嘗試指定
    對應的工作表),並恢復到上一個選取的分頁.
    """
    sel = event.widget.select()
    try:
        txt = event.widget.tab(sel, 'text') or ''
    except Exception:
        txt = ''
    if txt.upper() == 'EXCEL':
        # 由上一個被選取的 tab 判斷要開啟的 sheet(1-based index)
        sheet_idx = None
        try:
            last_id = _last_tab.get('id')
            if last_id:
                last_text = event.widget.tab(last_id, 'text') or ''
                lt = str(last_text).strip().lower()
                if 'rust' in lt:
                    sheet_idx = 1
                elif 'any' in lt:
                    sheet_idx = 2
                elif 'vnc' in lt or 'tight' in lt:
                    sheet_idx = 3
        except Exception:
            sheet_idx = None

        open_alldesk_excel(sheet_idx)
        try:
            # 還原到上一個 tab
            event.widget.select(_last_tab['id'])
        except Exception:
            pass
    else:
        _last_tab['id'] = sel

notebook.bind('<<NotebookTabChanged>>', _on_tab_changed)

# 將主視窗置中於螢幕
try:
    gui.update_idletasks()
    w = gui.winfo_width() or gui.winfo_reqwidth()
    h = gui.winfo_height() or gui.winfo_reqheight()
    sw = gui.winfo_screenwidth()
    sh = gui.winfo_screenheight()
    x = max((sw - w) // 2, 0)
    y = max((sh - h) // 2, 0)
    gui.geometry(f'+{x}+{y}')
except Exception:
    pass

gui.mainloop()